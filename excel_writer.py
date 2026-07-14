"""
excel_writer.py

Generates the three output Excel workbooks (Upload_File.xlsx,
Validation_Report.xlsx, Error_Report.xlsx) with consistent,
professional styling using openpyxl.
"""

from __future__ import annotations

from io import BytesIO
from pathlib import Path

import pandas as pd
from openpyxl import Workbook
from openpyxl.formatting.rule import FormulaRule
from openpyxl.styles import Alignment, Font, PatternFill
from openpyxl.utils import get_column_letter
from openpyxl.worksheet.worksheet import Worksheet

from config import (
    COL_VALIDATION_STATUS,
    ERROR_REPORT_FILE_NAME,
    UPLOAD_FILE_NAME,
    VALIDATION_REPORT_FILE_NAME,
    ValidationStatus,
)


class ExcelWriteError(Exception):
    """Raised when an output workbook cannot be written to disk."""


# --------------------------------------------------------------------------
# Styling constants
# --------------------------------------------------------------------------
_HEADER_FILL = PatternFill(start_color="1F3864", end_color="1F3864", fill_type="solid")
_HEADER_FONT = Font(color="FFFFFF", bold=True, size=11)
_HEADER_ALIGNMENT = Alignment(horizontal="center", vertical="center", wrap_text=True)

_ALT_ROW_FILL = PatternFill(start_color="F2F7FB", end_color="F2F7FB", fill_type="solid")

_STATUS_FILLS: dict[str, PatternFill] = {
    ValidationStatus.READY_FOR_UPLOAD.value: PatternFill(
        start_color="C6EFCE", end_color="C6EFCE", fill_type="solid"
    ),
    ValidationStatus.RRP_MISMATCH.value: PatternFill(
        start_color="FFEB9C", end_color="FFEB9C", fill_type="solid"
    ),
    ValidationStatus.SRP_MISMATCH.value: PatternFill(
        start_color="FFEB9C", end_color="FFEB9C", fill_type="solid"
    ),
    ValidationStatus.BOTH_MISMATCH.value: PatternFill(
        start_color="FFC7CE", end_color="FFC7CE", fill_type="solid"
    ),
    ValidationStatus.SKU_NOT_FOUND.value: PatternFill(
        start_color="FFC7CE", end_color="FFC7CE", fill_type="solid"
    ),
    ValidationStatus.DUPLICATE_SKU.value: PatternFill(
        start_color="D9D2E9", end_color="D9D2E9", fill_type="solid"
    ),
    ValidationStatus.MISSING_DATA.value: PatternFill(
        start_color="F4CCCC", end_color="F4CCCC", fill_type="solid"
    ),
    ValidationStatus.INVALID_PRICE.value: PatternFill(
        start_color="F4CCCC", end_color="F4CCCC", fill_type="solid"
    ),
}


class ExcelWriter:
    """Writes validation outputs to styled Excel workbooks."""

    def __init__(self, output_dir: Path) -> None:
        """
        Args:
            output_dir: Directory in which output workbooks are created.
        """
        self.output_dir = Path(output_dir)

    def _prepare_output_dir(self) -> None:
        try:
            self.output_dir.mkdir(parents=True, exist_ok=True)
        except OSError as exc:
            raise ExcelWriteError(
                f"Unable to create output directory '{self.output_dir}': {exc}"
            ) from exc

    def _style_worksheet(
        self,
        worksheet: Worksheet,
        df: pd.DataFrame,
        status_column_name: str | None = None,
    ) -> None:
        """Apply header styling, alternating row shading, autofilter,
        frozen panes, column auto-sizing, and optional status-based
        conditional fill to a worksheet already populated with df.

        Row coloring is implemented via native Excel conditional
        formatting rules (declarative, O(1) regardless of row count)
        rather than per-cell fills, so this remains fast for reports
        with 100,000+ rows.
        """

        n_rows = df.shape[0]
        n_cols = df.shape[1]

        # Header row styling
        for col_idx in range(1, n_cols + 1):
            cell = worksheet.cell(row=1, column=col_idx)
            cell.fill = _HEADER_FILL
            cell.font = _HEADER_FONT
            cell.alignment = _HEADER_ALIGNMENT

        if n_rows > 0:
            last_col_letter = get_column_letter(n_cols)
            data_range = f"A2:{last_col_letter}{n_rows + 1}"

            status_col_idx: int | None = None
            if status_column_name and status_column_name in df.columns:
                status_col_idx = list(df.columns).index(status_column_name) + 1

            if status_col_idx is not None:
                status_col_letter = get_column_letter(status_col_idx)
                # Status-specific rules take priority (added first) and
                # stop further (zebra) rules from overriding them.
                for status_value, fill in _STATUS_FILLS.items():
                    formula = f'${status_col_letter}2="{status_value}"'
                    worksheet.conditional_formatting.add(
                        data_range,
                        FormulaRule(formula=[formula], fill=fill, stopIfTrue=True),
                    )

            # Zebra banding for any row not already colored by a status rule.
            worksheet.conditional_formatting.add(
                data_range,
                FormulaRule(
                    formula=["MOD(ROW(),2)=0"], fill=_ALT_ROW_FILL, stopIfTrue=True
                ),
            )

        # Auto-size columns based on content length (vectorized per column)
        for col_idx in range(1, n_cols + 1):
            column_letter = get_column_letter(col_idx)
            header_len = len(str(df.columns[col_idx - 1]))
            if n_rows > 0:
                max_data_len = int(
                    df.iloc[:, col_idx - 1]
                    .map(lambda value: len(str(value)))
                    .max()
                )
            else:
                max_data_len = 0
            width = max(header_len, max_data_len) + 4
            worksheet.column_dimensions[column_letter].width = min(max(width, 12), 60)

        # Freeze header row, enable autofilter, hide gridlines
        worksheet.freeze_panes = "A2"
        if n_rows > 0:
            worksheet.auto_filter.ref = (
                f"A1:{get_column_letter(n_cols)}{n_rows + 1}"
            )
        worksheet.sheet_view.showGridLines = False

    def _build_workbook(
        self,
        df: pd.DataFrame,
        sheet_title: str,
        status_column_name: str | None = None,
    ) -> Workbook:
        """Build a styled openpyxl Workbook from a DataFrame (no I/O)."""
        workbook = Workbook()
        worksheet = workbook.active
        worksheet.title = sheet_title[:31]  # Excel sheet name limit

        columns = list(df.columns)
        worksheet.append(columns)

        for row in df.itertuples(index=False, name=None):
            worksheet.append(list(row))

        self._style_worksheet(worksheet, df, status_column_name)
        return workbook

    def _write_dataframe(
        self,
        df: pd.DataFrame,
        file_name: str,
        sheet_title: str,
        status_column_name: str | None = None,
    ) -> Path:
        self._prepare_output_dir()
        output_path = self.output_dir / file_name

        try:
            workbook = self._build_workbook(df, sheet_title, status_column_name)
            workbook.save(output_path)
        except OSError as exc:
            raise ExcelWriteError(
                f"Unable to write output file '{output_path}': {exc}"
            ) from exc
        except Exception as exc:  # noqa: BLE001 - surface any openpyxl failure
            raise ExcelWriteError(
                f"Failed to generate workbook '{output_path}': {exc}"
            ) from exc

        return output_path

    def _dataframe_to_bytes(
        self,
        df: pd.DataFrame,
        sheet_title: str,
        status_column_name: str | None = None,
    ) -> bytes:
        """
        Build a styled workbook entirely in memory and return its bytes.

        Used in environments with no persistent/writable filesystem
        (e.g. Streamlit Community Cloud), where output is delivered to
        the user via a download button instead of a saved file.
        """
        try:
            workbook = self._build_workbook(df, sheet_title, status_column_name)
            buffer = BytesIO()
            workbook.save(buffer)
            return buffer.getvalue()
        except Exception as exc:  # noqa: BLE001 - surface any openpyxl failure
            raise ExcelWriteError(
                f"Failed to generate workbook '{sheet_title}': {exc}"
            ) from exc

    def write_upload_file(self, upload_df: pd.DataFrame) -> Path:
        """Write the Upload_File.xlsx containing only Ready for Upload rows."""
        return self._write_dataframe(
            upload_df, UPLOAD_FILE_NAME, sheet_title="Upload File"
        )

    def write_validation_report(self, validation_report_df: pd.DataFrame) -> Path:
        """Write the Validation_Report.xlsx with full validation results."""
        return self._write_dataframe(
            validation_report_df,
            VALIDATION_REPORT_FILE_NAME,
            sheet_title="Validation Report",
            status_column_name=COL_VALIDATION_STATUS,
        )

    def write_error_report(self, error_report_df: pd.DataFrame) -> Path:
        """Write the Error_Report.xlsx listing all structural errors."""
        return self._write_dataframe(
            error_report_df, ERROR_REPORT_FILE_NAME, sheet_title="Error Report"
        )

    def upload_file_bytes(self, upload_df: pd.DataFrame) -> bytes:
        """Build Upload_File.xlsx in memory and return its bytes."""
        return self._dataframe_to_bytes(upload_df, sheet_title="Upload File")

    def validation_report_bytes(self, validation_report_df: pd.DataFrame) -> bytes:
        """Build Validation_Report.xlsx in memory and return its bytes."""
        return self._dataframe_to_bytes(
            validation_report_df,
            sheet_title="Validation Report",
            status_column_name=COL_VALIDATION_STATUS,
        )

    def error_report_bytes(self, error_report_df: pd.DataFrame) -> bytes:
        """Build Error_Report.xlsx in memory and return its bytes."""
        return self._dataframe_to_bytes(error_report_df, sheet_title="Error Report")
